"""
routes/batch.py — Batch pipeline page and all /api/batch/* endpoints.

Blueprint prefix: none (mounted at /)
Routes:
  GET  /batch                    → batch pipeline dashboard page
  GET  /api/batch/candidates     → list of all candidate DOIs + current status
  GET  /api/batch/result/<doi>   → full pipeline result dict for one DOI
  POST /api/batch/run            → start batch, stream SSE progress
  POST /api/batch/stop           → set stop flag on running batch
  POST /api/batch/export         → write export file; body: {"type": "review"|"all"|"minimal"|"selected", "format": "csv"|"xlsx"|"pdf", "dois": [...]}
  POST /api/batch/refresh        → reload source files and re-filter candidates
"""
import datetime
import json
import math
import queue
import threading
from pathlib import Path
from urllib.parse import unquote

import pandas as pd
from flask import Blueprint, Response, jsonify, render_template, request, stream_with_context

from validate import state
from shared.config import (
    CACHE_DIR, DATA_DIR,
    FILTERED_CSV_PATH, FINAL_OUTPUT_PATH, FLORA_SHEET_PATH,
    OPENALEX_CANDS_PATH, PDF_CACHE_DIR, REVIEW_CSV_PATH, log,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from fpdf import FPDF
from extract.link_original import run_for_doi
from shared.utils import cache_key, clean_doi, pdf_serve_url

VALIDATION_FILE = CACHE_DIR / "validations.json"


def _load_validations():
    if VALIDATION_FILE.exists():
        try:
            with VALIDATION_FILE.open(encoding="utf-8") as fh:
                state.validations.update(json.load(fh))
        except Exception as e:
            log.warning("Could not load validations: %s", e)


def _save_validations():
    try:
        with VALIDATION_FILE.open("w", encoding="utf-8") as fh:
            json.dump(state.validations, fh, ensure_ascii=False, indent=2)
    except Exception as e:
        log.warning("Could not save validations: %s", e)


_load_validations()

batch_bp = Blueprint("batch", __name__)

# ── Batch state (module-level, thread-safe via lock) ─────────────────────────
_batch_lock  = threading.Lock()
_stop_event  = threading.Event()
_batch_state = {
    "running"    : False,
    "current_doi": None,
    "total"      : 0,
    "done"       : 0,
}

# ── Review CSV column selection ───────────────────────────────────────────────
_REVIEW_COLS = [
    "doi_r", "study_r", "year_r",
    "resolved_doi_o", "resolved_title_o", "resolved_year_o", "resolved_author_o",
    "resolution_method", "resolution_score",
    "flora_doi_o", "flora_study_o", "flora_outcome",
    "flora_tag",
    "pdf_source", "pdf_url",
    "llm_source", "llm_confidence", "llm_evidence",
]

_MINIMAL_COLS = [
    "doi_r", "study_r", "resolved_doi_o", "resolved_title_o",
    "user_val_status", "flora_validation_status",
]


def _coerce_cell(value):
    """Convert a cell value to a safe string for Excel export, handling NaN."""
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return str(value)


_PDF_CHAR_MAP = str.maketrans({
    "—": "-",   # em dash
    "–": "-",   # en dash
    "‘": "'",   # left single quote
    "’": "'",   # right single quote
    "“": '"',   # left double quote
    "”": '"',   # right double quote
    "…": "...", # ellipsis
    " ": " ",   # non-breaking space
})

def _pdf_text(text: str) -> str:
    """Sanitize text for fpdf2 core fonts (Latin-1 only)."""
    text = text.translate(_PDF_CHAR_MAP)
    return text.encode("latin-1", errors="replace").decode("latin-1")


def _write_xlsx(df: pd.DataFrame, path) -> None:
    """Write df to an Excel file at path with a styled header row."""
    wb = Workbook()
    ws = wb.active

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")

    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill

    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=_coerce_cell(value))

    for col_idx, col_name in enumerate(df.columns, 1):
        values = df[col_name].astype(str).tolist()
        max_len = max(len(col_name), *(len(v) for v in values), 1)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    wb.save(str(path))


def _write_pdf(df: pd.DataFrame, title: str, path) -> None:
    """Write df to a PDF table at path. Landscape for wide tables."""
    n_cols = len(df.columns)
    orientation = "L" if n_cols > 6 else "P"
    font_size   = 5 if n_cols > 17 else (7 if n_cols > 6 else 9)

    pdf = FPDF(orientation=orientation, unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=10)
    pdf.add_page()

    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(0, 8, _pdf_text(title), new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", "B", font_size)

    usable_w = pdf.w - pdf.l_margin - pdf.r_margin
    col_w    = usable_w / n_cols
    row_h    = font_size * 0.6

    # Header row
    pdf.set_fill_color(30, 58, 95)
    pdf.set_text_color(255, 255, 255)
    for col_name in df.columns:
        pdf.cell(col_w, row_h + 1, _pdf_text(str(col_name)[:25]), border=1, fill=True)
    pdf.ln()

    # Data rows
    pdf.set_font("Helvetica", "", font_size)
    pdf.set_text_color(0, 0, 0)
    chars_per_cell = max(int(col_w / (font_size * 0.18)), 30)

    for i, row in enumerate(df.itertuples(index=False)):
        if i % 2 == 0:
            pdf.set_fill_color(240, 248, 255)
        else:
            pdf.set_fill_color(255, 255, 255)
        for value in row:
            text = _pdf_text(str(value or "")[:chars_per_cell])
            pdf.cell(col_w, row_h, text, border=1, fill=True)
        pdf.ln()

    pdf.output(str(path))


def _write_export(export_type: str, fmt: str, dois: list) -> dict:
    """
    Build and write an export file to DATA_DIR.

    export_type : "review" | "all" | "minimal" | "selected"
    fmt         : "csv" | "xlsx" | "pdf"
    dois        : DOI list — used to filter rows for "minimal" and "selected"

    Returns {"rows": N, "filename": "flora_review.xlsx", "path": "..."} or
            {"error": "..."}.
    """
    with state.resolved_lock:
        if not state.resolved:
            return {"error": "no resolved results yet"}

        rows = []
        for doi_r, result in state.resolved.items():
            v = state.validations.get(doi_r, {})
            rows.append({
                **result,
                "doi_r"           : doi_r,
                "flora_tag"       : _flora_tag(result),
                "user_val_status" : v.get("status",  ""),
                "user_val_comment": v.get("comment", ""),
            })

    full_df = pd.DataFrame(rows)

    # Row filter (minimal / selected use the DOI list from the browser)
    if export_type == "selected":
        if not dois:
            return {"error": "dois required for selected export"}
        doi_set = {clean_doi(d) for d in dois}
        full_df = full_df[full_df["doi_r"].apply(clean_doi).isin(doi_set)]
    elif export_type == "minimal" and dois:
        doi_set = {clean_doi(d) for d in dois}
        full_df = full_df[full_df["doi_r"].apply(clean_doi).isin(doi_set)]

    # Column filter
    if export_type == "review":
        cols = [c for c in _REVIEW_COLS if c in full_df.columns]
    elif export_type == "all":
        cols = list(full_df.columns)
    else:  # minimal, selected
        cols = [c for c in _MINIMAL_COLS if c in full_df.columns]

    export_df = full_df[cols] if cols else full_df

    # Output path
    ext_map  = {"csv": ".csv", "xlsx": ".xlsx", "pdf": ".pdf"}
    filename = f"flora_{export_type}{ext_map[fmt]}"
    path     = DATA_DIR / filename

    try:
        if fmt == "csv":
            export_df.to_csv(path, index=False, encoding="utf-8-sig")
        elif fmt == "xlsx":
            _write_xlsx(export_df, path)
        elif fmt == "pdf":
            n     = len(export_df)
            title = f"FLoRA Export — {export_type} ({n} row{'s' if n != 1 else ''})"
            _write_pdf(export_df, title, path)
    except PermissionError:
        return {
            "error": f"{filename} is open in another application. "
                     f"Close the file and try again."
        }

    return {"rows": len(export_df), "filename": filename, "path": str(path)}


# ── Helpers ───────────────────────────────────────────────────────────────────

def _flora_tag(result: dict) -> str:
    """Compute the FLoRA comparison tag string for a result dict."""
    res_doi   = (result.get("resolved_doi_o")   or "").lower()
    res_title = (result.get("resolved_title_o") or "").lower().strip()
    if not res_doi and not res_title:
        return "unresolved"
    flora_doi   = (result.get("flora_doi_o")   or "").lower()
    flora_title = (result.get("flora_study_o") or "").lower().strip()
    if not flora_doi and not flora_title:
        return "new"
    if res_doi == flora_doi or res_title == flora_title:
        return "confirms_flora"
    return "updated_from_flora"


def _row_status(doi_r: str) -> str:
    """Return display status for a DOI based on current resolved dict."""
    if doi_r == _batch_state.get("current_doi"):
        return "running"
    r = state.resolved.get(doi_r)
    if r is None:
        return "pending"
    return "resolved" if (r.get("resolved_doi_o") or r.get("resolved_title_o")) else "unresolved"


def _upsert_resolved_csv_locked(doi_r: str, result: dict) -> None:
    """Write/update a single row in multiple_match_resolved.csv. Caller must hold resolved_lock."""
    row = {**result, "doi_r": doi_r, "flora_tag": _flora_tag(result)}
    new_df = pd.DataFrame([row])
    if FINAL_OUTPUT_PATH.exists():
        existing = pd.read_csv(FINAL_OUTPUT_PATH, dtype=str, encoding="utf-8-sig").fillna("")
        existing["doi_r"] = existing["doi_r"].apply(clean_doi)
        existing = existing[existing["doi_r"] != clean_doi(doi_r)]
        combined = pd.concat([existing, new_df], ignore_index=True)
    else:
        combined = new_df
    combined.to_csv(FINAL_OUTPUT_PATH, index=False, encoding="utf-8-sig")


def _build_candidate_row(row: pd.Series) -> dict:
    """Build the candidate dict for a single filtered_df row."""
    doi_r  = clean_doi(str(row.get("doi_r", "")))
    result = state.resolved.get(doi_r) or {}
    status = _row_status(doi_r)
    # validation_status: prefer resolved result (may be refreshed), fall back to CSV row
    validation_status = (
        result.get("validation_status")
        or str(row.get("validation_status", "") or "")
    )
    flora_validation_status = (
        result.get("flora_validation_status")
        or str(row.get("flora_validation_status", "") or "")
    )
    v = state.validations.get(doi_r, {})
    return {
        "doi_r"                  : doi_r,
        "study_r"                : str(row.get("study_r", "")),
        "year_r"                 : str(row.get("year_r", "")),
        "status"                 : status,
        "resolution_method"      : result.get("resolution_method", ""),
        "resolved_title_o"       : result.get("resolved_title_o", ""),
        "resolved_doi_o"         : result.get("resolved_doi_o",   ""),
        "flora_tag"              : _flora_tag(result) if result else "",
        "validation_status"      : validation_status,
        "flora_validation_status": flora_validation_status,
        "user_val_status"        : v.get("status",  ""),
        "user_val_comment"       : v.get("comment", ""),
        "pdf_serve_url"     : pdf_serve_url(doi_r, result) if result else "",
        "pdf_ok"            : bool(result.get("pdf_ok", False)) if result else False,
        "pdf_url"           : result.get("pdf_url",    "") if result else "",
        "pdf_source"        : result.get("pdf_source", "") if result else "",
        "llm_source"        : result.get("llm_source", "") if result else "",
        "has_result"        : bool(result),
    }


# ── Page routes ───────────────────────────────────────────────────────────────

@batch_bp.route("/batch")
def batch_page():
    return render_template("batch.html", active_page="batch")


# ── API: dashboard stats ──────────────────────────────────────────────────────

@batch_bp.route("/api/batch/stats")
def api_dashboard_stats():
    """Return summary statistics for the dashboard page."""
    total_candidates = len(state.filtered_df) if not state.filtered_df.empty else 0

    with state.resolved_lock:
        results = dict(state.resolved)

    processed   = len(results)
    pending     = max(0, total_candidates - processed)

    pdf_cached  = 0
    pdf_url_only = 0
    pdf_none    = 0

    resolved_count   = 0
    unresolved_count = 0

    method_counts : dict[str, int] = {}
    flora_tag_counts: dict[str, int] = {}
    llm_conf_counts : dict[str, int] = {}

    for doi_r, r in results.items():
        is_resolved = bool(r.get("resolved_doi_o") or r.get("resolved_title_o"))
        if is_resolved:
            resolved_count += 1
        else:
            unresolved_count += 1

        pdf_ok  = r.get("pdf_ok")
        if isinstance(pdf_ok, str):
            pdf_ok = pdf_ok.lower() in ("true", "1", "yes")
        pdf_path = r.get("pdf_path", "")
        pdf_url  = r.get("pdf_url", "")

        has_cached = bool(pdf_ok)
        if not has_cached and pdf_path:
            has_cached = Path(pdf_path).exists()
        if not has_cached:
            expected = PDF_CACHE_DIR / f"{cache_key(doi_r)}.pdf"
            has_cached = expected.exists()

        if has_cached:
            pdf_cached += 1
        elif pdf_url:
            pdf_url_only += 1
        else:
            pdf_none += 1

        meth = r.get("resolution_method") or "none"
        method_counts[meth] = method_counts.get(meth, 0) + 1

        tag = _flora_tag(r)
        flora_tag_counts[tag] = flora_tag_counts.get(tag, 0) + 1

        conf = r.get("llm_confidence") or ""
        if conf:
            llm_conf_counts[conf] = llm_conf_counts.get(conf, 0) + 1

    return jsonify({
        "total_candidates": total_candidates,
        "processed"       : processed,
        "pending"         : pending,
        "resolved"        : resolved_count,
        "unresolved"      : unresolved_count,
        "pdf_cached"      : pdf_cached,
        "pdf_url_only"    : pdf_url_only,
        "pdf_none"        : pdf_none,
        "method_counts"   : method_counts,
        "flora_tag_counts": flora_tag_counts,
        "llm_conf_counts" : llm_conf_counts,
    })


# ── API: candidates list ──────────────────────────────────────────────────────

@batch_bp.route("/api/batch/candidates")
def api_candidates():
    """Return all candidate DOIs with their current resolution status."""
    if state.filtered_df.empty:
        return jsonify([])
    rows = [_build_candidate_row(row) for _, row in state.filtered_df.iterrows()]
    return jsonify(rows)


# ── API: single result ────────────────────────────────────────────────────────

@batch_bp.route("/api/batch/result/<path:doi>")
def api_result(doi: str):
    doi_r  = clean_doi(unquote(doi))
    result = state.resolved.get(doi_r)
    if result is None:
        return jsonify({"error": "not found"}), 404
    return jsonify(result)


# ── API: export ───────────────────────────────────────────────────────────────

@batch_bp.route("/api/batch/export", methods=["POST"])
def api_export():
    body        = request.get_json(force=True) or {}
    export_type = body.get("type",   "review")
    fmt         = body.get("format", "csv")
    dois        = body.get("dois",   [])

    if export_type not in ("review", "all", "minimal", "selected"):
        return jsonify({"error": f"unknown type: {export_type}"}), 400
    if fmt not in ("csv", "xlsx", "pdf"):
        return jsonify({"error": f"unknown format: {fmt}"}), 400

    result = _write_export(export_type, fmt, dois)
    if "error" in result:
        return jsonify(result), 400
    return jsonify(result)


# ── API: refresh candidates ───────────────────────────────────────────────────

@batch_bp.route("/api/batch/refresh", methods=["POST"])
def api_refresh():
    """Reload source data from disk and re-filter candidates."""
    def _load(path, label):
        if not path.exists():
            log.warning("%s not found at %s", label, path)
            return pd.DataFrame()
        df = pd.read_csv(path, dtype=str, encoding="utf-8-sig").fillna("")
        log.info("Refreshed %s: %d rows", label, len(df))
        return df

    state.flora_df = _load(FLORA_SHEET_PATH,    "FLoRA entry sheet")
    state.cands_df = _load(OPENALEX_CANDS_PATH, "openalex_candidates")

    if "doi_r" in state.flora_df.columns:
        state.flora_df["doi_r"] = state.flora_df["doi_r"].apply(clean_doi)
    if "doi_r" in state.cands_df.columns:
        state.cands_df["doi_r"] = state.cands_df["doi_r"].apply(clean_doi)

    if state.flora_df.empty or state.cands_df.empty:
        return jsonify({"error": "source data not found", "count": 0})

    if "prep_notes" in state.flora_df.columns:
        flora_dois = set(
            state.flora_df[state.flora_df["prep_notes"].str.startswith("openalex:", na=False)]["doi_r"]
        ) - {""}
    else:
        flora_dois = set()

    multi_mask = (
        state.cands_df["doi_r"].isin(flora_dois)
        & (state.cands_df["match_source"] == "openalex_references")
        & (state.cands_df["match_status"] == "multiple_matches")
    )
    state.filtered_df = state.cands_df[multi_mask].copy().reset_index(drop=True)

    if not state.filtered_df.empty:
        state.filtered_df.to_csv(FILTERED_CSV_PATH, index=False, encoding="utf-8-sig")

    log.info("Refresh complete: %d candidates", len(state.filtered_df))
    return jsonify({"count": len(state.filtered_df)})


# ── API: stop ─────────────────────────────────────────────────────────────────

@batch_bp.route("/api/batch/stop", methods=["POST"])
def api_stop():
    with _batch_lock:
        if not _batch_state["running"]:
            return jsonify({"status": "not_running"})
        _stop_event.set()
    return jsonify({"status": "stopping"})


# ── API: run (SSE) ────────────────────────────────────────────────────────────

@batch_bp.route("/api/batch/run", methods=["POST"])
def api_run():
    """
    Start the batch pipeline. Streams Server-Sent Events as each DOI completes.

    Request body:
      {"mode": "all" | "unresolved" | "selected" | "smart", "dois": [...]}

    SSE event types:
      {"type": "progress", "doi": "...", "index": 4, "total": 11, "result": {...}}
      {"type": "error",    "doi": "...", "index": 4, "total": 11, "error": "..."}
      {"type": "done",     "total": 11, "resolved": 8}
      {"type": "stopped"}
    """
    body          = request.get_json(force=True) or {}
    mode          = body.get("mode", "all")
    selected_dois = [clean_doi(d) for d in body.get("dois", []) if d]
    force_rerun   = bool(body.get("force", False))

    with _batch_lock:
        if _batch_state["running"]:
            return jsonify({"error": "batch already running"}), 409
        if state.filtered_df.empty:
            return jsonify({"error": "no candidates loaded"}), 400

        all_dois = [clean_doi(str(r["doi_r"])) for _, r in state.filtered_df.iterrows()]

        if mode == "selected":
            queue_list = [d for d in selected_dois if d in set(all_dois)]
        elif mode == "unresolved":
            queue_list = [
                d for d in all_dois
                if not (state.resolved.get(d) or {}).get("resolved_doi_o")
                and not (state.resolved.get(d) or {}).get("resolved_title_o")
            ]
        elif mode == "smart":
            queue_list = [
                d for d in all_dois
                if state.validations.get(d, {}).get("status") != "successful"
            ]
        else:
            queue_list = all_dois

        if not queue_list:
            return jsonify({"error": "no DOIs to process"}), 400

        _stop_event.clear()
        _batch_state["running"]     = True
        _batch_state["current_doi"] = None
        _batch_state["total"]       = len(queue_list)
        _batch_state["done"]        = 0

    result_queue: queue.Queue = queue.Queue()

    def worker():
        for i, doi_r in enumerate(queue_list):
            if _stop_event.is_set():
                result_queue.put({"type": "stopped"})
                break

            _batch_state["current_doi"] = doi_r
            _batch_state["done"]        = i + 1

            try:
                result = run_for_doi(doi_r, state.flora_df, state.cands_df,
                                     force=force_rerun)
                result["pdf_serve_url"] = pdf_serve_url(doi_r, result)
                result["flora_tag"]     = _flora_tag(result)

                with state.resolved_lock:
                    state.resolved[doi_r] = result
                    _upsert_resolved_csv_locked(doi_r, result)

                result_queue.put({
                    "type"  : "progress",
                    "doi"   : doi_r,
                    "index" : i + 1,
                    "total" : len(queue_list),
                    "result": result,
                })
            except Exception as e:
                log.exception("Batch pipeline error for %s", doi_r)
                result_queue.put({
                    "type" : "error",
                    "doi"  : doi_r,
                    "index": i + 1,
                    "total": len(queue_list),
                    "error": str(e),
                })

        if not _stop_event.is_set():
            resolved_count = sum(
                1 for r in state.resolved.values()
                if r.get("resolved_doi_o") or r.get("resolved_title_o")
            )
            result_queue.put({
                "type"    : "done",
                "total"   : len(queue_list),
                "resolved": resolved_count,
            })

        with _batch_lock:
            _batch_state["running"]     = False
            _batch_state["current_doi"] = None

    thread = threading.Thread(target=worker, daemon=True)
    thread.start()

    def generate():
        while True:
            try:
                event = result_queue.get(timeout=120)
            except queue.Empty:
                yield "data: {\"type\": \"heartbeat\"}\n\n"
                continue
            yield f"data: {json.dumps(event, ensure_ascii=False)}\n\n"
            if event["type"] in ("done", "stopped"):
                break

    return Response(
        stream_with_context(generate()),
        content_type="text/event-stream",
        headers={
            "Cache-Control"    : "no-cache",
            "X-Accel-Buffering": "no",
            "Connection"       : "keep-alive",
        },
    )


# ── API: human validation ─────────────────────────────────────────────────────

@batch_bp.route("/api/batch/validate", methods=["POST"])
def api_validate():
    """Save a human validation status + comment for a DOI."""
    body   = request.get_json(force=True) or {}
    doi_r  = clean_doi(body.get("doi_r", ""))
    status = body.get("validation_status", "")
    comment = body.get("comment", "")
    if not doi_r or status not in ("successful", "failed", "recheck"):
        return jsonify({"error": "invalid input"}), 400
    state.validations[doi_r] = {
        "status"   : status,
        "comment"  : comment,
        "timestamp": datetime.datetime.utcnow().isoformat(),
    }
    _save_validations()
    return jsonify({"ok": True})


# ── API: single-DOI re-run with validation comment ────────────────────────────

@batch_bp.route("/api/batch/run_doi", methods=["POST"])
def api_run_doi():
    """Re-run the pipeline for one DOI with an optional validator note."""
    body    = request.get_json(force=True) or {}
    doi_r   = clean_doi(body.get("doi_r", ""))
    comment = body.get("validation_comment", "")
    if not doi_r:
        return jsonify({"error": "doi_r required"}), 400
    try:
        result = run_for_doi(doi_r, state.flora_df, state.cands_df,
                             force=True, validation_comment=comment)
        result["pdf_serve_url"] = pdf_serve_url(doi_r, result)
        result["flora_tag"]     = _flora_tag(result)
        with state.resolved_lock:
            state.resolved[doi_r] = result
            _upsert_resolved_csv_locked(doi_r, result)
        return jsonify(result)
    except Exception as e:
        log.exception("run_doi error for %s", doi_r)
        return jsonify({"error": str(e)}), 500
